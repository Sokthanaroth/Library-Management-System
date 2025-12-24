# library_app/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.urls import reverse_lazy
from django.http import JsonResponse
from django.db.models import Count

from .models import Book, Member, BorrowRecord, Category, Author, Publisher, Location
from django.contrib.auth.models import User
from .forms import (
    BookForm, CategoryForm, MemberForm, UserRegistrationForm,
    UserUpdateForm, MemberUpdateForm, AuthorForm, PublisherForm,
    BorrowRecordForm
)
from django.db.models import Q
from .decorators import staff_required, admin_required

# ---------------------- Custom Checks ----------------------
def librarian_required(user):
    return user.is_authenticated and user.is_staff

# ---------------------- PUBLIC VIEWS ----------------------
def home(request):
    books = Book.objects.all().order_by('-id')[:8]
    return render(request, 'library_app/home.html', {'books': books})

def book_list(request):
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    category_filter = request.GET.get('category', '')
    year_filter = request.GET.get('year', '')

    books = Book.objects.select_related('publisher').prefetch_related('categories', 'authors').all()

    if query:
        books = books.filter(
            Q(title__icontains=query) |
            Q(authors__name__icontains=query) |
            Q(isbn__icontains=query)
        ).distinct()
    if category_filter:
        books = books.filter(categories__id=category_filter)
    if status_filter:
        books = books.filter(status=status_filter.lower())
    if year_filter:
        books = books.filter(published_date__year=year_filter)

    paginator = Paginator(books, 12)
    page_number = request.GET.get('page')
    try:
        books = paginator.page(page_number)
    except PageNotAnInteger:
        books = paginator.page(1)
    except EmptyPage:
        books = paginator.page(paginator.num_pages)

    categories = Category.objects.filter(is_active=True).order_by('name')
    status_choices = Book.STATUS_CHOICES

    context = {
        'books': books,
        'query': query,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'year_filter': year_filter,
        'categories': categories,
        'status_choices': status_choices,
    }
    return render(request, 'library_app/book_list.html', context)

def book_detail(request, pk):
    book = get_object_or_404(Book.objects.select_related('publisher').prefetch_related('categories', 'authors'), pk=pk)
    return render(request, 'library_app/book_detail.html', {'book': book})

def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            Member.objects.create(user=user, member_id=f"M{user.id:04d}")
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password1')
            from django.contrib.auth import authenticate, login
            user = authenticate(username=username, password=password)
            login(request, user)
            messages.success(request, 'Registration successful!')
            return redirect('home')
        else:
            messages.error(request, f"Please correct the errors: {form.errors}")
    else:
        form = UserRegistrationForm()
    return render(request, 'library_app/register.html', {'form': form})

# ---------------------- DASHBOARD ----------------------
@login_required
@user_passes_test(librarian_required)
def dashboard(request):
    # Basic statistics
    total_books = Book.objects.count()
    total_members = Member.objects.count()
    borrowed_books = BorrowRecord.objects.filter(return_date__isnull=True).count()
    overdue_books = BorrowRecord.objects.filter(return_date__isnull=True, due_date__lt=timezone.now()).count()

    # Recent activities (last 10 activities)
    recent_activities = []

    # Recent borrowings (last 5)
    recent_borrows = BorrowRecord.objects.select_related('book', 'member__user').order_by('-borrow_date')[:5]
    for borrow in recent_borrows:
        recent_activities.append({
            'type': 'borrow',
            'icon': 'bi-arrow-up-circle',
            'color': 'success',
            'title': f'Book borrowed: {borrow.book.title}',
            'description': f'by {borrow.member.user.get_full_name() or borrow.member.user.username}',
            'timestamp': borrow.borrow_date,
            'time_ago': _get_time_ago(borrow.borrow_date)
        })

    # Recent returns (last 5)
    recent_returns = BorrowRecord.objects.select_related('book', 'member__user').filter(
        return_date__isnull=False
    ).order_by('-return_date')[:5]
    for return_record in recent_returns:
        fine_text = f" (Fine: ${return_record.fine_amount:.2f})" if return_record.fine_amount > 0 else ""
        recent_activities.append({
            'type': 'return',
            'icon': 'bi-arrow-down-circle',
            'color': 'info',
            'title': f'Book returned: {return_record.book.title}',
            'description': f'by {return_record.member.user.get_full_name() or return_record.member.user.username}{fine_text}',
            'timestamp': return_record.return_date,
            'time_ago': _get_time_ago(return_record.return_date)
        })

    # Recent book additions (last 3)
    recent_books = Book.objects.select_related('created_by').order_by('-created_at')[:3]
    for book in recent_books:
        recent_activities.append({
            'type': 'book_added',
            'icon': 'bi-plus-circle',
            'color': 'primary',
            'title': f'New book added: {book.title}',
            'description': f'by {book.created_by.get_full_name() or book.created_by.username}',
            'timestamp': book.created_at,
            'time_ago': _get_time_ago(book.created_at)
        })

    # Sort activities by timestamp (most recent first)
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:10]  # Keep only the 10 most recent

    # Additional dashboard data
    popular_books = Book.objects.annotate(
        borrow_count=Count('borrows')
    ).order_by('-borrow_count')[:5]

    low_stock_books = Book.objects.filter(available_copies__lte=2, available_copies__gt=0)[:5]

    context = {
        'total_books': total_books,
        'total_members': total_members,
        'borrowed_books': borrowed_books,
        'overdue_books': overdue_books,
        'recent_activities': recent_activities,
        'popular_books': popular_books,
        'low_stock_books': low_stock_books,
    }
    return render(request, 'library_app/dashboard.html', context)

def _get_time_ago(timestamp):
    """Helper function to get human-readable time ago"""
    now = timezone.now()
    diff = now - timestamp

    if diff.days > 0:
        return f"{diff.days} day{'s' if diff.days != 1 else ''} ago"
    elif diff.seconds >= 3600:
        hours = diff.seconds // 3600
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    elif diff.seconds >= 60:
        minutes = diff.seconds // 60
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    else:
        return "Just now"


# ---------------------- BOOK CRUD (Staff/Admin) ----------------------
@login_required
@staff_required
def book_create(request):
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES)
        if form.is_valid():
            book = form.save(commit=False)

            # Handle selected publisher
            selected_publisher = form.cleaned_data.get('publisher')
            if selected_publisher:
                book.publisher = selected_publisher

            # Handle new publisher (overrides selected if entered)
            new_publisher_name = form.cleaned_data.get('new_publisher')
            if new_publisher_name:
                publisher, _ = Publisher.objects.get_or_create(name=new_publisher_name)
                book.publisher = publisher

            # Handle location
            loc_name = form.cleaned_data.get('location_name')
            if loc_name:
                location, _ = Location.objects.get_or_create(name=loc_name)
                book.location = location

            book.created_by = request.user
            book.save()

            # Handle selected authors
            selected_authors = form.cleaned_data.get('authors')
            if selected_authors:
                book.authors.set(selected_authors)

            # Handle new authors
            new_authors = form.cleaned_data.get('new_authors')
            if new_authors:
                author_names = [a.strip() for a in new_authors.split(',') if a.strip()]
                for name in author_names:
                    author, _ = Author.objects.get_or_create(name=name)
                    book.authors.add(author)

            # Handle categories (ManyToMany field needs explicit saving)
            selected_categories = form.cleaned_data.get('categories')
            if selected_categories:
                book.categories.set(selected_categories)
            else:
                book.categories.clear()

            messages.success(request, 'Book added successfully!')
            return redirect('book_list')
        else:
            messages.error(request, f"Please correct the errors: {form.errors}")
    else:
        form = BookForm()
    return render(request, 'library_app/book_form.html', {'form': form, 'title': 'Add Book'})


@login_required
@staff_required
def book_update(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == 'POST':
        form = BookForm(request.POST, request.FILES, instance=book)
        if form.is_valid():
            book = form.save(commit=False)

            # Handle selected publisher
            selected_publisher = form.cleaned_data.get('publisher')
            if selected_publisher:
                book.publisher = selected_publisher

            # Handle new publisher (overrides selected if entered)
            new_publisher_name = form.cleaned_data.get('new_publisher')
            if new_publisher_name:
                publisher, _ = Publisher.objects.get_or_create(name=new_publisher_name)
                book.publisher = publisher

            loc_name = form.cleaned_data.get('location_name')
            if loc_name:
                location, _ = Location.objects.get_or_create(name=loc_name)
                book.location = location
            book.save()

            # Handle selected authors
            selected_authors = form.cleaned_data.get('authors')
            if selected_authors:
                book.authors.set(selected_authors)
            else:
                book.authors.clear()

            # Handle new authors
            new_authors = form.cleaned_data.get('new_authors')
            if new_authors:
                author_names = [a.strip() for a in new_authors.split(',') if a.strip()]
                for name in author_names:
                    author, _ = Author.objects.get_or_create(name=name)
                    book.authors.add(author)

            # Handle categories (ManyToMany field needs explicit saving)
            selected_categories = form.cleaned_data.get('categories')
            if selected_categories:
                book.categories.set(selected_categories)
            else:
                book.categories.clear()

            messages.success(request, 'Book updated successfully!')
            return redirect('book_detail', pk=book.pk)
        else:
            messages.error(request, f"Please correct the errors: {form.errors}")
    else:
        form = BookForm(instance=book)
    return render(request, 'library_app/book_form.html', {'form': form, 'title': 'Edit Book'})


@login_required
@staff_required
def book_delete(request, pk):
    book = get_object_or_404(Book, pk=pk)
    if request.method == 'POST':
        book.delete()
        messages.success(request, 'Book deleted successfully.')
        return redirect('book_list')
    return render(request, 'library_app/book_confirm_delete.html', {'book': book})


# ---------------------- BORROW & RETURN (Staff/Admin) ----------------------
@login_required
@staff_required
def borrow_book(request, book_id):
    book = get_object_or_404(Book, pk=book_id)

    if not book.is_available:
        messages.error(request, "Book is not available for borrowing.")
        return redirect('book_detail', pk=book_id)

    if request.method == 'POST':
        form = BorrowRecordForm(request.POST)
        if form.is_valid():
            try:
                member = form.cleaned_data['member']
                borrow_record = book.borrow(member)
                # Update the borrow record with the form data
                borrow_record.due_date = form.cleaned_data['due_date']
                borrow_record.save()
                messages.success(request, f'Book "{book.title}" borrowed successfully by {member.user.get_full_name() or member.user.username}.')
                return redirect('borrow_records')
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('borrow_form', book_id=book_id)  # Keep form data on error
        else:
            messages.error(request, "Please correct the form errors.")
    else:
        form = BorrowRecordForm()

    return render(request, 'library_app/borrow_form.html', {'book': book, 'form': form})




@login_required
@staff_required
def return_book(request, borrow_id):
    borrow = get_object_or_404(BorrowRecord.objects.select_related('book', 'member'), pk=borrow_id)

    if request.method == 'POST':
        try:
            borrow.book.return_book(borrow)
            fine_amount = borrow.fine_amount
            member = borrow.member

            # Automatically deduct fine from member's balance
            if fine_amount > 0:
                member.fine_balance -= fine_amount
                member.save()
                messages.warning(request, f'Book "{borrow.book.title}" returned successfully. Fine of ${fine_amount:.2f} has been automatically deducted from your balance.')
            else:
                messages.success(request, f'Book "{borrow.book.title}" returned successfully.')
            return redirect('borrow_records')
        except ValueError as e:
            messages.error(request, str(e))
            return redirect('return_book', borrow_id=borrow_id)

    return render(request, 'library_app/return_book.html', {'borrow': borrow})


# ---------------------- MEMBER MANAGEMENT (Staff/Admin) ----------------------
@login_required
@staff_required
def member_list(request):
    # Staff can access member management for creating member cards and managing balances
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    role_filter = request.GET.get('role', '')

    members = Member.objects.select_related('user').all()

    if query:
        members = members.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__username__icontains=query) |
            Q(member_id__icontains=query)
        )

    if status_filter:
        if status_filter == 'active':
            members = members.filter(is_active=True)
        elif status_filter == 'inactive':
            members = members.filter(is_active=False)

    if role_filter:
        members = members.filter(role=role_filter)

    # Add pagination
    paginator = Paginator(members, 12)  # Show 12 members per page
    page_number = request.GET.get('page')
    try:
        members = paginator.page(page_number)
    except PageNotAnInteger:
        members = paginator.page(1)
    except EmptyPage:
        members = paginator.page(paginator.num_pages)

    context = {
        'members': members,
        'query': query,
        'status_filter': status_filter,
        'role_filter': role_filter,
    }
    return render(request, 'library_app/member_list.html', context)


@login_required
@staff_required
def member_topup(request, member_id):
    member = get_object_or_404(Member, pk=member_id)

    # ABA payment link
    aba_payment_link = "https://link.payway.com.kh/aba?id=58355315E753&dynamic=true&source_caller=sdk&pid=af_app_invites&link_action=abaqr&shortlink=n5eoyxxs&created_from_app=true&acc=500113895&af_siteid=968860649&userid=58355315E753&code=060691&c=abaqr&af_referrer_uid=1612001769086-2722499"

    if request.method == 'POST':
        action = request.POST.get('action', 'generate_qr')

        if action == 'generate_qr':
            amount_str = request.POST.get('amount', '').strip()
            try:
                # Handle comma as decimal separator (common in some locales)
                amount_str = amount_str.replace(',', '.')
                amount = float(amount_str)
                if amount > 0:
                    try:
                        # Generate QR code for ABA payment
                        import qrcode
                        import base64
                        from io import BytesIO

                        # Append amount to payment link
                        payment_link_with_amount = f"{aba_payment_link}&amount={amount:.2f}"

                        qr = qrcode.QRCode(version=1, box_size=10, border=5)
                        qr.add_data(payment_link_with_amount)
                        qr.make(fit=True)

                        img = qr.make_image(fill='black', back_color='white')
                        buffer = BytesIO()
                        img.save(buffer, format='PNG')
                        buffer.seek(0)
                        qr_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

                        context = {
                            'member': member,
                            'amount': amount,
                            'qr_code': f"data:image/png;base64,{qr_base64}",
                            'payment_link': payment_link_with_amount,
                            'show_qr': True
                        }
                        return render(request, 'library_app/member_topup.html', context)
                    except Exception as e:
                        messages.error(request, f"Error generating QR code: {str(e)}")
                else:
                    messages.error(request, "Amount must be positive.")
            except (ValueError, TypeError):
                messages.error(request, "Please enter a valid amount (e.g., 25.50)")

        elif action == 'confirm_payment':
            # After payment confirmation, add the balance
            amount_str = request.POST.get('confirmed_amount', '').strip()
            try:
                amount = float(amount_str)
                if amount > 0:
                    from decimal import Decimal
                    amount_decimal = Decimal(str(amount))
                    member.fine_balance += amount_decimal
                    try:
                        member.save()
                        messages.success(request, f"Top-up of ${amount:.2f} processed for {member.user.get_full_name() or member.user.username}.")
                        return redirect('member_detail', pk=member.pk)
                    except Exception as e:
                        messages.error(request, f"Error saving balance update: {str(e)}")
                else:
                    messages.error(request, "Invalid amount.")
            except (ValueError, TypeError):
                messages.error(request, "Invalid amount.")

    return render(request, 'library_app/member_topup.html', {'member': member})


# ---------------------- USER & MEMBER CREATION ----------------------
@login_required
@admin_required
def user_create(request):
    """Superuser can create users with passwords"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('user_create')

        if not password:
            messages.error(request, "Password is required.")
            return redirect('user_create')

        # Create user with password
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name,
            last_name=last_name,
            is_staff=is_staff,
            is_superuser=is_superuser
        )

        # Create member automatically
        Member.objects.create(user=user, member_id=f"M{user.id:04d}")

        messages.success(request, f"User '{username}' created successfully.")
        return redirect('member_list')

    return render(request, 'library_app/user_create.html')

@login_required
@staff_required
def member_create(request):
    """Staff can create members without passwords - automatically sets as member role"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone = request.POST.get('phone')
        address = request.POST.get('address')

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists.")
            return redirect('member_create_staff')

        # Create user without password
        user = User.objects.create_user(
            username=username,
            email=email,
            first_name=first_name,
            last_name=last_name
        )

        # Create member with automatic 'member' role (not guest)
        member = Member.objects.create(
            user=user,
            member_id=f"M{user.id:04d}",
            phone=phone,
            address=address,
            role='member'  # Automatically set as member, not guest
        )

        messages.success(request, f"Member '{username}' created successfully.")
        return redirect('member_list')

    return render(request, 'library_app/member_create.html')

# ---------------------- PROFILE ----------------------
@login_required
def profile_view(request):
    user = request.user
    member, _ = Member.objects.get_or_create(user=user, defaults={'member_id': f"M{user.id:04d}"})

    # Get borrowing history
    borrow_records = BorrowRecord.objects.filter(member=member).select_related('book').order_by('-borrow_date')

    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=user)
        member_form = MemberUpdateForm(request.POST, instance=member)
        if user_form.is_valid() and member_form.is_valid():
            user_form.save()
            member_form.save()
            messages.success(request, "Profile updated successfully!")
            return redirect('profile')
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        user_form = UserUpdateForm(instance=user)
        member_form = MemberUpdateForm(instance=member)

    return render(request, 'library_app/profile.html', {
        'user_form': user_form,
        'member_form': member_form,
        'member': member,
        'borrow_records': borrow_records,
    })


# ---------------------- CATEGORY CRUD ----------------------
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

class CategoryListView(ListView):
    model = Category
    template_name = 'library_app/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        queryset = Category.objects.filter(is_active=True)

        # Get filter parameters
        query = self.request.GET.get('q', '').strip()
        color_filter = self.request.GET.get('color', '').strip()
        min_books = self.request.GET.get('min_books', '').strip()

        # Apply filters
        if query:
            queryset = queryset.filter(name__icontains=query)

        if color_filter and color_filter != '#000000':
            queryset = queryset.filter(color=color_filter)

        if min_books:
            try:
                min_books_int = int(min_books)
                if min_books_int > 0:
                    # Filter categories that have at least min_books_int books
                    queryset = queryset.annotate(
                        books_count=Count('books')
                    ).filter(books_count__gte=min_books_int)
            except ValueError:
                pass  # Ignore invalid min_books values

        return queryset.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Pass filter values to template for form persistence
        context['query'] = self.request.GET.get('q', '')
        context['color_filter'] = self.request.GET.get('color', '#000000')
        context['min_books'] = self.request.GET.get('min_books', '')
        return context

class CategoryCreateView(CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'library_app/category_form.html'
    success_url = reverse_lazy('category_list')

class CategoryUpdateView(UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'library_app/category_form.html'
    success_url = reverse_lazy('category_list')

class CategoryDeleteView(DeleteView):
    model = Category
    template_name = 'library_app/category_confirm_delete.html'
    success_url = reverse_lazy('category_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.is_active = False
        self.object.save()
        return redirect('category_list')

def get_categories_json(request):
    categories = Category.objects.filter(is_active=True).values('id', 'name', 'color')
    return JsonResponse(list(categories), safe=False)

# ---------------------- CATEGORY DETAIL VIEW ----------------------
def category_detail(request, pk):
    """Display books in a specific category"""
    category = get_object_or_404(Category, pk=pk, is_active=True)

    # Get books in this category with pagination
    books = Book.objects.filter(categories=category).select_related('publisher').prefetch_related('authors', 'categories').order_by('title')

    # Pagination
    paginator = Paginator(books, 12)
    page_number = request.GET.get('page')
    try:
        books = paginator.page(page_number)
    except PageNotAnInteger:
        books = paginator.page(1)
    except EmptyPage:
        books = paginator.page(paginator.num_pages)

    context = {
        'category': category,
        'books': books,
    }
    return render(request, 'library_app/category_detail.html', context)

# ---------------------- AUTHOR CRUD ----------------------
@login_required
@staff_required
def author_list(request):
    authors = Author.objects.all().order_by('name')
    query = request.GET.get('q', '')
    if query:
        authors = authors.filter(name__icontains=query)
    return render(request, 'library_app/author_list.html', {'authors': authors, 'query': query})

@login_required
@staff_required
def author_detail(request, pk):
    author = get_object_or_404(Author, pk=pk)
    return render(request, 'library_app/author_detail.html', {'author': author})

@login_required
@staff_required
def author_create(request):
    if request.method == 'POST':
        form = AuthorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Author added successfully!')
            return redirect('author_list')
    else:
        form = AuthorForm()
    return render(request, 'library_app/author_form.html', {'form': form, 'title': 'Add Author'})

@login_required
@staff_required
def author_update(request, pk):
    author = get_object_or_404(Author, pk=pk)
    if request.method == 'POST':
        form = AuthorForm(request.POST, instance=author)
        if form.is_valid():
            form.save()
            messages.success(request, 'Author updated successfully!')
            return redirect('author_detail', pk=author.pk)
    else:
        form = AuthorForm(instance=author)
    return render(request, 'library_app/author_form.html', {'form': form, 'title': 'Edit Author'})

@login_required
@staff_required
def author_delete(request, pk):
    author = get_object_or_404(Author, pk=pk)
    if request.method == 'POST':
        author.delete()
        messages.success(request, 'Author deleted successfully!')
        return redirect('author_list')
    return render(request, 'library_app/author_confirm_delete.html', {'author': author})

# ---------------------- PUBLISHER CRUD ----------------------
@login_required
@staff_required
def publisher_list(request):
    query = request.GET.get('q', '')
    publishers = Publisher.objects.all().order_by('name')
    if query:
        publishers = publishers.filter(name__icontains=query)
    return render(request, 'library_app/publisher_list.html', {'publishers': publishers, 'query': query})

@login_required
@staff_required
def publisher_detail(request, pk):
    publisher = get_object_or_404(Publisher, pk=pk)
    return render(request, 'library_app/publisher_detail.html', {'publisher': publisher})

@login_required
@staff_required
def publisher_create(request):
    if request.method == 'POST':
        form = PublisherForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Publisher added successfully!')
            return redirect('publisher_list')
    else:
        form = PublisherForm()
    return render(request, 'library_app/publisher_form.html', {'form': form, 'title': 'Add Publisher'})

@login_required
@staff_required
def publisher_update(request, pk):
    publisher = get_object_or_404(Publisher, pk=pk)
    if request.method == 'POST':
        form = PublisherForm(request.POST, instance=publisher)
        if form.is_valid():
            form.save()
            messages.success(request, 'Publisher updated successfully!')
            return redirect('publisher_detail', pk=publisher.pk)
    else:
        form = PublisherForm(instance=publisher)
    return render(request, 'library_app/publisher_form.html', {'form': form, 'title': 'Edit Publisher'})

@login_required
@staff_required
def publisher_delete(request, pk):
    publisher = get_object_or_404(Publisher, pk=pk)
    if request.method == 'POST':
        publisher.delete()
        messages.success(request, 'Publisher deleted successfully!')
        return redirect('publisher_list')
    return render(request, 'library_app/publisher_confirm_delete.html', {'publisher': publisher})
@login_required
@staff_required
def member_detail(request, pk):
    member = get_object_or_404(Member, pk=pk)
    return render(request, 'library_app/member_detail.html', {'member': member})
@login_required
@staff_required
def member_edit(request, pk):
    member = get_object_or_404(Member, pk=pk)
    # Use the UserUpdateForm and MemberForm so the template receives form objects
    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=member.user)
        member_form = MemberForm(request.POST, instance=member)

        if user_form.is_valid() and member_form.is_valid():
            # Enforce role permissions: only superusers may assign 'staff' or 'admin'
            requested_role = member_form.cleaned_data.get('role')
            if requested_role in ['staff', 'admin'] and not request.user.is_superuser:
                messages.error(request, "Access denied. Only admins can assign staff/admin roles.")
                return redirect('member_edit', pk=pk)

            # Prevent non-superusers from changing role of existing staff/admin members
            if member.role in ['staff', 'admin'] and not request.user.is_superuser:
                messages.error(request, "Access denied. Only admins can change roles of staff/admin users.")
                return redirect('member_edit', pk=pk)

            # Save user and member
            user_form.save()
            member = member_form.save()

            # Update User permissions based on role
            user = member.user
            if member.role in ['staff', 'admin']:
                user.is_staff = True
                if member.role == 'admin':
                    user.is_superuser = True
            else:
                user.is_staff = False
                # Note: is_superuser is not unset to avoid demoting existing superusers
            user.save()

            messages.success(request, 'Member updated successfully!')
            return redirect('member_detail', pk=member.pk)
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        user_form = UserUpdateForm(instance=member.user)
        member_form = MemberForm(instance=member)

    return render(request, 'library_app/member_form.html', {
        'user_form': user_form,
        'member_form': member_form,
        'member': member,
        'title': 'Edit Member',
    })

@login_required
@staff_required
def member_delete(request, pk):
    # Only superusers (admins) can delete members - staff cannot delete
    if not request.user.is_superuser:
        messages.error(request, "Access denied. Only admins can delete members.")
        return redirect('member_list')

    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        member.user.delete()
        member.delete()
        messages.success(request, 'Member deleted successfully!')
        return redirect('member_list')
    return render(request, 'library_app/member_confirm_delete.html', {'member': member})

# ---------------------- BARCODE & REPORTS ----------------------
@login_required
@staff_required
def generate_barcode(request, book_id=None, member_id=None):
    """Generate and display barcode for a book or member library card"""
    if book_id:
        obj = get_object_or_404(Book, pk=book_id)
        obj_type = 'book'
    elif member_id:
        obj = get_object_or_404(Member, pk=member_id)
        obj_type = 'member'
    else:
        messages.error(request, "Invalid request")
        return redirect('home')

    try:
        if obj_type == 'book':
            barcode_image = obj.generate_barcode_image()
            title = obj.title
            barcode_value = obj.barcode
        else:
            # For members, we'll use their member_id as barcode
            barcode_value = obj.member_id
            title = f"Library Card - {obj.user.get_full_name() or obj.user.username}"

            # Generate barcode for member
            import barcode
            from barcode.writer import ImageWriter
            from io import BytesIO

            code128 = barcode.get_barcode_class('code128')
            barcode_instance = code128(barcode_value, writer=ImageWriter())

            options = {
                'write_text': False,
                'module_height': 15.0,
                'module_width': 0.5,
                'quiet_zone': 2.0,
            }

            buffer = BytesIO()
            barcode_instance.write(buffer, options=options)
            buffer.seek(0)
            barcode_image = buffer.getvalue()

        # Convert to base64 for display in template
        import base64
        barcode_base64 = base64.b64encode(barcode_image).decode('utf-8')

        context = {
            'obj': obj,
            'obj_type': obj_type,
            'title': title,
            'barcode_value': barcode_value,
            'barcode_image': f"data:image/png;base64,{barcode_base64}",
        }
        return render(request, 'library_app/barcode_display.html', context)
    except Exception as e:
        messages.error(request, f"Error generating barcode: {str(e)}")
        if obj_type == 'book':
            return redirect('book_detail', pk=book_id)
        else:
            return redirect('member_detail', pk=member_id)

@login_required
@staff_required
def print_barcodes(request):
    """Print barcodes for selected books"""
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from io import BytesIO

    # Get selected book IDs from POST data
    book_ids = request.POST.getlist('book_ids', [])

    if not book_ids:
        messages.error(request, "No books selected for barcode printing.")
        return redirect('book_list')

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="book_barcodes.pdf"'

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    # Get books
    books = Book.objects.filter(id__in=book_ids)

    # Layout settings
    x_start = 0.5 * inch
    y_start = 10.5 * inch
    label_width = 2.5 * inch
    label_height = 1.5 * inch
    margin = 0.2 * inch

    x = x_start
    y = y_start

    for book in books:
        try:
            # Generate barcode
            barcode_image = book.generate_barcode_image()

            # Save barcode to temporary file
            import tempfile
            import os
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_file:
                tmp_file.write(barcode_image)
                tmp_filename = tmp_file.name

            # Draw book info and barcode
            c.setFont("Helvetica-Bold", 10)
            c.drawString(x + margin, y - 0.2 * inch, book.title[:25])

            c.setFont("Helvetica", 8)
            c.drawString(x + margin, y - 0.4 * inch, f"ISBN: {book.isbn}")

            # Draw barcode image
            c.drawImage(tmp_filename, x + margin, y - 1.2 * inch, width=2 * inch, height=0.8 * inch)

            # Clean up temp file
            os.unlink(tmp_filename)

            # Move to next position
            x += label_width + margin
            if x > 7 * inch:  # New row
                x = x_start
                y -= label_height + margin
                if y < 1 * inch:  # New page
                    c.showPage()
                    y = y_start

        except Exception as e:
            # Skip books that can't generate barcodes
            continue

    c.save()
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()

    return response

@login_required
@staff_required
def generate_report(request):
    """Generate library reports in Excel format"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    report_type = request.GET.get('type', 'books')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="library_report_{report_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    if report_type == 'books':
        # Books report
        books = Book.objects.select_related('publisher').prefetch_related('authors', 'categories').all()

        # Headers
        headers = ['Title', 'Subtitle', 'ISBN', 'Authors', 'Publisher', 'Categories', 'Status', 'Available Copies', 'Published Date', 'Edition', 'Page Count']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data
        for row_num, book in enumerate(books, 2):
            ws.cell(row=row_num, column=1, value=book.title).border = border
            ws.cell(row=row_num, column=2, value=book.subtitle or '').border = border
            ws.cell(row=row_num, column=3, value=book.isbn).border = border
            ws.cell(row=row_num, column=4, value=book.author_names).border = border
            ws.cell(row=row_num, column=5, value=book.publisher.name if book.publisher else '').border = border
            ws.cell(row=row_num, column=6, value=book.category_names).border = border
            ws.cell(row=row_num, column=7, value=book.get_status_display()).border = border
            ws.cell(row=row_num, column=8, value=book.available_copies).border = border
            ws.cell(row=row_num, column=9, value=book.published_date.strftime('%Y-%m-%d') if book.published_date else '').border = border
            ws.cell(row=row_num, column=10, value=book.edition).border = border
            ws.cell(row=row_num, column=11, value=book.page_count or '').border = border

    elif report_type == 'borrow_records':
        # Borrow records report
        records = BorrowRecord.objects.select_related('book', 'member__user').order_by('-borrow_date')

        # Headers
        headers = ['Book Title', 'ISBN', 'Member Name', 'Member ID', 'Borrow Date', 'Due Date', 'Return Date', 'Status', 'Fine Amount', 'Days Overdue']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data
        for row_num, record in enumerate(records, 2):
            status = 'Returned' if record.return_date else ('Overdue' if record.is_overdue else 'Active')
            days_overdue = (timezone.now().date() - record.due_date).days if record.is_overdue else 0

            ws.cell(row=row_num, column=1, value=record.book.title).border = border
            ws.cell(row=row_num, column=2, value=record.book.isbn).border = border
            ws.cell(row=row_num, column=3, value=record.member.user.get_full_name() or record.member.user.username).border = border
            ws.cell(row=row_num, column=4, value=record.member.member_id).border = border
            ws.cell(row=row_num, column=5, value=record.borrow_date.strftime('%Y-%m-%d %H:%M')).border = border
            ws.cell(row=row_num, column=6, value=record.due_date.strftime('%Y-%m-%d') if record.due_date else '').border = border
            ws.cell(row=row_num, column=7, value=record.return_date.strftime('%Y-%m-%d %H:%M') if record.return_date else '').border = border
            ws.cell(row=row_num, column=8, value=status).border = border
            ws.cell(row=row_num, column=9, value=float(record.fine_amount)).border = border
            ws.cell(row=row_num, column=10, value=days_overdue).border = border

    elif report_type == 'members':
        # Members report
        members = Member.objects.select_related('user').all()

        # Headers
        headers = ['Member ID', 'Username', 'Full Name', 'Email', 'Phone', 'Address', 'Role', 'Membership Date', 'Expiration Date', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data
        for row_num, member in enumerate(members, 2):
            ws.cell(row=row_num, column=1, value=member.member_id).border = border
            ws.cell(row=row_num, column=2, value=member.user.username).border = border
            ws.cell(row=row_num, column=3, value=member.user.get_full_name()).border = border
            ws.cell(row=row_num, column=4, value=member.user.email).border = border
            ws.cell(row=row_num, column=5, value=member.phone or '').border = border
            ws.cell(row=row_num, column=6, value=member.address or '').border = border
            ws.cell(row=row_num, column=7, value=member.get_role_display()).border = border
            ws.cell(row=row_num, column=8, value=member.membership_date.strftime('%Y-%m-%d') if member.membership_date else '').border = border
            ws.cell(row=row_num, column=9, value=member.expiration_date.strftime('%Y-%m-%d') if member.expiration_date else '').border = border
            ws.cell(row=row_num, column=10, value='Active' if member.is_membership_valid else 'Expired').border = border

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response

# ---------------------- LIBRARY CARD PDF ----------------------
@login_required
@staff_required
def generate_library_card(request, member_id):
    """Generate and download library card PDF for a member"""
    from django.http import FileResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import credit_card
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    import barcode
    from barcode.writer import ImageWriter
    from io import BytesIO

    member = get_object_or_404(Member, pk=member_id)

    buffer = BytesIO()
    # Credit card size: 85.6mm x 54mm
    c = canvas.Canvas(buffer, pagesize=(85.6*mm, 54*mm))

    # Draw background (you can replace with your image path)
    # c.drawImage("path/to/your/background.png", 0, 0, width=85.6*mm, height=54*mm)

    # Draw library title
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(1, 0.55, 0)  # Orange
    c.drawString(10*mm, 45*mm, "Library Card")

    # Draw organization name
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0, 0, 0)
    c.drawString(60*mm, 50*mm, "My Library Org")

    # Draw member name
    c.setFont("Helvetica-Bold", 9)
    c.drawString(10*mm, 40*mm, f"Name: {member.user.get_full_name() or member.user.username}")

    # Draw card number
    c.setFont("Helvetica", 8)
    c.drawString(10*mm, 35*mm, f"Card #: {member.card_number}")

    # Generate barcode image
    CODE128 = barcode.get_barcode_class('code128')
    barcode_img = BytesIO()
    CODE128(str(member.card_number), writer=ImageWriter()).write(barcode_img)
    barcode_img.seek(0)
    barcode_reader = ImageReader(barcode_img)
    c.drawImage(barcode_reader, 10*mm, 5*mm, width=60*mm, height=15*mm)

    c.showPage()
    c.save()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename=f'library_card_{member.user.username}.pdf')

# ---------------------- BORROW RECORDS MANAGEMENT ----------------------
@login_required
@staff_required
def barcode_scanner(request):
    """Barcode scanner interface for quick operations"""
    return render(request, 'library_app/barcode_scanner.html')

@login_required
@staff_required
def scan_barcode(request):
    """Handle barcode scanning for various operations"""
    from django.http import JsonResponse

    if request.method == 'POST':
        barcode = request.POST.get('barcode', '').strip()
        action = request.POST.get('action', 'search')  # search, borrow, return

        if not barcode:
            return JsonResponse({'success': False, 'message': 'Barcode is required'})

        # Try to find book by barcode or ISBN
        book = None
        try:
            # First try exact barcode match
            book = Book.objects.filter(barcode=barcode).first()
            if not book:
                # Then try ISBN match
                book = Book.objects.filter(isbn=barcode).first()
        except:
            pass

        if not book:
            return JsonResponse({'success': False, 'message': 'Book not found with this barcode'})

        if action == 'search':
            return JsonResponse({
                'success': True,
                'action': 'search',
                'book': {
                    'id': book.id,
                    'title': book.title,
                    'author_names': book.author_names,
                    'isbn': book.isbn,
                    'status': book.get_status_display(),
                    'available_copies': book.available_copies,
                    'barcode': book.barcode,
                }
            })

        elif action == 'borrow':
            # Check if user is logged in and is a member
            if not hasattr(request.user, 'member'):
                return JsonResponse({'success': False, 'message': 'User is not a registered member'})

            member = request.user.member

            if not book.is_available:
                return JsonResponse({'success': False, 'message': 'Book is not available for borrowing'})

            try:
                borrow_record = book.borrow(member)
                return JsonResponse({
                    'success': True,
                    'action': 'borrow',
                    'message': f'Book "{book.title}" borrowed successfully',
                    'book': {
                        'id': book.id,
                        'title': book.title,
                        'available_copies': book.available_copies,
                        'status': book.get_status_display(),
                    }
                })
            except ValueError as e:
                return JsonResponse({'success': False, 'message': str(e)})

        elif action == 'return':
            # Find active borrow record for this book and current user
            try:
                borrow_record = BorrowRecord.objects.filter(
                    book=book,
                    member__user=request.user,
                    return_date__isnull=True
                ).first()

                if not borrow_record:
                    return JsonResponse({'success': False, 'message': 'No active borrow record found for this book'})

                book.return_book(borrow_record)
                fine_amount = borrow_record.fine_amount

                message = f'Book "{book.title}" returned successfully'
                if fine_amount > 0:
                    message += f'. Fine: ${fine_amount:.2f}'

                return JsonResponse({
                    'success': True,
                    'action': 'return',
                    'message': message,
                    'book': {
                        'id': book.id,
                        'title': book.title,
                        'available_copies': book.available_copies,
                        'status': book.get_status_display(),
                    },
                    'fine': float(fine_amount)
                })
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Error returning book: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Invalid request'})

@login_required
@staff_required
def bulk_scan(request):
    """Handle bulk barcode scanning operations"""
    from django.http import JsonResponse

    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            action = data.get('action', 'inventory')
            barcodes = data.get('items', [])

            if not barcodes:
                return JsonResponse({'success': False, 'message': 'No items to process'})

            results = []
            success_count = 0
            error_count = 0

            if action == 'inventory':
                # Inventory check - just verify books exist
                for barcode in barcodes:
                    book = None
                    try:
                        book = Book.objects.filter(Q(barcode=barcode) | Q(isbn=barcode)).first()
                        if book:
                            results.append({
                                'barcode': barcode,
                                'success': True,
                                'message': f'Found: {book.title}',
                                'book': {
                                    'id': book.id,
                                    'title': book.title,
                                    'status': book.get_status_display(),
                                    'available_copies': book.available_copies
                                }
                            })
                            success_count += 1
                        else:
                            results.append({
                                'barcode': barcode,
                                'success': False,
                                'message': 'Book not found in system'
                            })
                            error_count += 1
                    except Exception as e:
                        results.append({
                            'barcode': barcode,
                            'success': False,
                            'message': f'Error: {str(e)}'
                        })
                        error_count += 1

            elif action == 'bulk_return':
                # Bulk return - return all scanned books for current user
                member = request.user.member
                for barcode in barcodes:
                    try:
                        book = Book.objects.filter(Q(barcode=barcode) | Q(isbn=barcode)).first()
                        if not book:
                            results.append({
                                'barcode': barcode,
                                'success': False,
                                'message': 'Book not found'
                            })
                            error_count += 1
                            continue

                        # Find active borrow record
                        borrow_record = BorrowRecord.objects.filter(
                            book=book,
                            member=member,
                            return_date__isnull=True
                        ).first()

                        if borrow_record:
                            book.return_book(borrow_record)
                            fine_amount = borrow_record.fine_amount
                            member = borrow_record.member

                            # Automatically deduct fine from member's balance
                            if fine_amount > 0:
                                member.fine_balance -= fine_amount
                                member.save()

                            message = f'Returned: {book.title}'
                            if fine_amount > 0:
                                message += f' (Fine: ${fine_amount:.2f} automatically deducted from balance)'
                            results.append({
                                'barcode': barcode,
                                'success': True,
                                'message': message
                            })
                            success_count += 1
                        else:
                            results.append({
                                'barcode': barcode,
                                'success': False,
                                'message': 'No active borrow record found'
                            })
                            error_count += 1
                    except Exception as e:
                        results.append({
                            'barcode': barcode,
                            'success': False,
                            'message': f'Error: {str(e)}'
                        })
                        error_count += 1

            summary = f"Processed {len(barcodes)} items. Success: {success_count}, Errors: {error_count}"

            return JsonResponse({
                'success': True,
                'results': results,
                'summary': summary,
                'stats': {
                    'total': len(barcodes),
                    'success': success_count,
                    'errors': error_count
                }
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Processing error: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Invalid request'})

@login_required
@staff_required
def borrow_records_list(request):
    borrow_records = BorrowRecord.objects.select_related('book', 'member__user').order_by('-borrow_date')

    # Filter options
    status_filter = request.GET.get('status', '')
    member_filter = request.GET.get('member', '')
    overdue_filter = request.GET.get('overdue', '')

    if status_filter:
        if status_filter == 'active':
            borrow_records = borrow_records.filter(return_date__isnull=True)
        elif status_filter == 'returned':
            borrow_records = borrow_records.filter(return_date__isnull=False)

    if member_filter:
        borrow_records = borrow_records.filter(member__user__username__icontains=member_filter)

    if overdue_filter == 'yes':
        borrow_records = borrow_records.filter(return_date__isnull=True, due_date__lt=timezone.now())

    paginator = Paginator(borrow_records, 20)
    page_number = request.GET.get('page')
    try:
        borrow_records = paginator.page(page_number)
    except PageNotAnInteger:
        borrow_records = paginator.page(1)
    except EmptyPage:
        borrow_records = paginator.page(paginator.num_pages)

    context = {
        'borrow_records': borrow_records,
        'status_filter': status_filter,
        'member_filter': member_filter,
        'overdue_filter': overdue_filter,
    }
    return render(request, 'library_app/borrow_records.html', context)
